"""
excel 处理包
"""

from openpyxl import Workbook


class DownloadTool:
    @classmethod
    def simple_download(cls, column_data, table_data):
        wb = Workbook()
        sheet = wb.active
        sheet.append(column_data)
        for line_v in table_data:
            sheet.append(line_v)

        return wb

    @classmethod
    def table_children_download(cls, columndata, tabledata):
        def deep_first_work(v_dict, ret_list):
            ret_list.append([v_dict.get(_) for _ in field_list])
            if "children" in v_dict:
                for sub_v_dict in v_dict["children"]:
                    deep_first_work(sub_v_dict, ret_list)

        name_list, field_list = [], []
        ret_list = []
        for _ in columndata:
            name_list.append(_["label"])
            field_list.append(_["prop"])
        for v_dict in tabledata:
            deep_first_work(v_dict, ret_list)

        wb = Workbook()
        sheet = wb.active
        sheet.append(name_list)
        for line in ret_list:
            sheet.append(line)
        return wb

    @classmethod
    def header_merge_download(cls, columndata, tabledata):
        wb = Workbook()
        sheet = wb.active
        column_key_list = []
        start_t = 'A'
        for column in columndata:  # 两行
            if "children" in column:
                length = len(column["children"]) - 1
                end_c = cls._cell_index_trans(start_t, length)
                sheet.merge_cells(start_t + "1:" + end_c + "1")
                sheet[start_t + "1"] = column["label"]
                for num, col_c in enumerate(column["children"], 0):
                    alpha_ = cls._cell_index_trans(start_t, num)
                    sheet[alpha_ + "2"] = col_c["label"]
                    column_key_list.append(col_c["prop"])
                start_t = cls._cell_index_trans(start_t, len(column["children"]))
            else:
                c_range = start_t + "1:" + start_t + "2"
                sheet.merge_cells(c_range)
                sheet[start_t + "1"] = column["label"]
                column_key_list.append(column["prop"])
                start_t = cls._cell_index_trans(start_t, 1)
        # 表格数据
        for line in tabledata:
            sheet.append([line[key] for key in column_key_list])
            if "children" in line:
                for c_line in line["children"]:
                    sheet.append([c_line[key] for key in column_key_list])

        return wb

    @classmethod
    def _cell_index_trans(cls, start_alpha, size):
        alpha_list = list(start_alpha)
        alpha_list.reverse()
        nums = 0
        for i, _ in enumerate(alpha_list):
            init_a = (ord(_) - 65) + i * 26
            nums += init_a
        new_nums = nums + size
        out_alpha = ""
        flag = False
        while new_nums > 0:
            k = new_nums // 26
            if k:
                out_alpha += chr(k - 1 + 65)
                k = new_nums % 26
                if k == 0:
                    k = 1
                    flag = True
            else:
                if flag:
                    v = ((new_nums - 1) % 26)
                else:
                    v = (new_nums % 26)
                out_alpha += chr(v + 65)

            new_nums = k

        return out_alpha
